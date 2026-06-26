import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash

load_dotenv()

from app import app, db, Usuario

EXCEL_ENTRADA = 'TALLERES -  INSCRIPCIONES - PRESENCIAL (2).xlsx'
EXCEL_IMPORTADOS = 'reporte_importados.xlsx'
EXCEL_PENDIENTES = 'reporte_pendientes.xlsx'


def parsear_dni(raw):
    try:
        return str(int(float(str(raw)))) if raw and str(raw).strip() not in ('None', '', '-') else None
    except Exception:
        return None


def estilo_cabecera(ws, cols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col, titulo in enumerate(cols, start=1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")


def main():
    wb = openpyxl.load_workbook(EXCEL_ENTRADA)
    ws = wb.active

    importados = []
    pendientes = []

    with app.app_context():
        for row in ws.iter_rows(min_row=2, values_only=True):
            email_raw = row[0]
            nombre = str(row[1]).strip() if row[1] else ''
            apellido = str(row[2]).strip() if row[2] else ''
            raw_dni = row[3]

            tiene_algo = any(v is not None and str(v).strip() not in ('', 'None') for v in row[:4])
            if not tiene_algo:
                continue

            email = str(email_raw).strip().lower() if email_raw else None
            dni = parsear_dni(raw_dni)
            nombre_completo = f"{nombre} {apellido}".strip()

            if not email or email == 'none':
                pendientes.append({'email': '', 'nombre': nombre_completo, 'dni': dni or '', 'motivo': 'Sin email'})
                continue

            if not dni:
                pendientes.append({'email': email, 'nombre': nombre_completo, 'dni': '', 'motivo': 'Sin DNI'})
                continue

            password_hash = generate_password_hash(dni)
            with db.session.no_autoflush:
                usuario = Usuario.query.filter_by(email=email).first()
            if usuario:
                usuario.password_hash = password_hash
                usuario.modalidad = 'presencial'
                if nombre_completo:
                    usuario.nombre_completo = nombre_completo
                estado = 'Actualizado'
            else:
                usuario = Usuario(
                    email=email,
                    password_hash=password_hash,
                    nombre_completo=nombre_completo or email,
                    modalidad='presencial',
                )
                db.session.add(usuario)
                estado = 'Creado'

            importados.append({'email': email, 'nombre': nombre_completo, 'dni': dni, 'estado': estado})

        db.session.commit()

    # Excel 1: importados exitosamente
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Importados"
    cols1 = ['Email', 'Nombre completo', 'DNI', 'Modalidad', 'Estado']
    estilo_cabecera(ws1, cols1)
    for r in importados:
        ws1.append([r['email'], r['nombre'], r['dni'], 'presencial', r['estado']])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 30
    wb1.save(EXCEL_IMPORTADOS)

    # Excel 2: pendientes (sin DNI o sin email)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Pendientes"
    cols2 = ['Email', 'Nombre completo', 'DNI disponible', 'Motivo']
    estilo_cabecera(ws2, cols2)
    for r in pendientes:
        ws2.append([r['email'], r['nombre'], r['dni'], r['motivo']])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 35
    wb2.save(EXCEL_PENDIENTES)

    print(f"\n=== Importación completada ===")
    print(f"  Importados: {len(importados)} ({sum(1 for r in importados if r['estado']=='Creado')} creados, {sum(1 for r in importados if r['estado']=='Actualizado')} actualizados)")
    print(f"  Pendientes: {len(pendientes)}")
    print(f"\nArchivos generados:")
    print(f"  {EXCEL_IMPORTADOS}")
    print(f"  {EXCEL_PENDIENTES}")


if __name__ == '__main__':
    main()
