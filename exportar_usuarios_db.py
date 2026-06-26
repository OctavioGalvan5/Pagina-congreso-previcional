import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

from app import app, db, Usuario

EXCEL_SALIDA = 'usuarios_db.xlsx'


def estilo_cabecera(ws, cols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    for col, titulo in enumerate(cols, start=1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")


def main():
    with app.app_context():
        usuarios = Usuario.query.order_by(Usuario.nombre_completo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    cols = ['ID', 'Email', 'Nombre completo', 'Modalidad']
    estilo_cabecera(ws, cols)

    for u in usuarios:
        ws.append([u.id, u.email, u.nombre_completo, u.modalidad])

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15

    wb.save(EXCEL_SALIDA)
    print(f"Exportados {len(usuarios)} usuarios -> {EXCEL_SALIDA}")


if __name__ == '__main__':
    main()
